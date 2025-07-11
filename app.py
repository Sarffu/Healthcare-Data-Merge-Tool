import pandas as pd
from tkinter import Tk, filedialog, ttk, messagebox
import tkinter as tk
import openpyxl
from openpyxl.styles import PatternFill
import threading
import os
from PIL import Image, ImageTk # Pillow library for image handling
import warnings # Import warnings module

# Global variables
file1_path_global = None
file2_path_global = None
save_path_global = None
merged_df_global = None # This will store the DataFrame with 'Was_Updated' column
file1_sheet_name_global = None
file2_sheet_name_global = None

# Global variables for icons
schedule_icon = None
roaster_icon = None

# Suppress the specific openpyxl UserWarning for invalid dates.
# This makes the application less noisy in the console for known data issues.
# The underlying data problem in the Excel file still exists, but our code
# handles it gracefully by coercing invalid dates to NaT.
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl.worksheet._reader')


def get_excel_sheet_names(file_path):
    """Returns a list of sheet names from an Excel file."""
    try:
        xls = pd.ExcelFile(file_path)
        return xls.sheet_names
    except Exception as e:
        raise ValueError(f"Error reading sheets from Excel file '{os.path.basename(file_path)}': {e}")

# Helper function to read Excel or CSV based on file extension
# Enhanced to handle date parsing more robustly after reading
def read_file_into_df(file_path, sheet_name=None):
    try:
        df = None
        if file_path.lower().endswith(('.xlsx', '.xls')):
            # Read Excel without parsing dates directly, let Pandas handle it later
            df = pd.read_excel(file_path, sheet_name=sheet_name, keep_default_na=True)
        elif file_path.lower().endswith('.csv'):
            # Detect separator for CSV
            try:
                df = pd.read_csv(file_path, keep_default_na=True)
                # Check if it looks like a single column (meaning wrong delimiter)
                if len(df.columns) == 1 and ';' in str(df.iloc[0,0]): # Check first cell for semicolon
                    df = pd.read_csv(file_path, sep=';', keep_default_na=True)
            except Exception:
                # Fallback if initial read fails, try common delimiters
                try:
                    df = pd.read_csv(file_path, sep=',', keep_default_na=True)
                except Exception:
                    df = pd.read_csv(file_path, sep=';', keep_default_na=True) # Try semicolon
        else:
            raise ValueError("Unsupported file format. Please select an Excel (.xlsx, .xls) or CSV (.csv) file.")
        
        # Strip whitespace from column names immediately after reading
        df.columns = df.columns.str.strip()

        # Explicitly convert potential date columns to datetime objects
        # using errors='coerce' to turn invalid parses into NaT
        if 'VotedDate' in df.columns:
            df['VotedDate'] = pd.to_datetime(df['VotedDate'], errors='coerce').dt.date # Added .dt.date
        if 'Provider Effective Date' in df.columns:
            df['Provider Effective Date'] = pd.to_datetime(df['Provider Effective Date'], errors='coerce').dt.date # Added .dt.date
            
        return df

    except Exception as e:
        raise ValueError(f"Error reading file '{os.path.basename(file_path)}': {e}")

def select_file_and_sheet(file_type):
    global file1_path_global, file2_path_global, file1_sheet_name_global, file2_sheet_name_global

    file_path = filedialog.askopenfilename(
        title=f"Import {file_type} File",
        filetypes=[("All Supported Files", "*.xlsx *.xls *.csv"),
                   ("Excel files", "*.xlsx *.xls"),
                   ("CSV files", "*.csv"),
                   ("All files", "*.*")],
    )

    if not file_path:
        if file_type == "Scheduler":
            file1_status.config(text="No file selected", foreground="red")
            file1_path_global = None
            file1_sheet_name_global = None
        elif file_type == "Roaster":
            file2_status.config(text="No file selected", foreground="red")
            file2_path_global = None
            file2_sheet_name_global = None
        check_and_enable_merge_button()
        return

    if file_path.lower().endswith(('.xlsx', '.xls')):
        try:
            sheet_names = get_excel_sheet_names(file_path)
            if not sheet_names:
                messagebox.showerror("Error", "No sheets found in the Excel file.")
                if file_type == "Scheduler":
                    file1_path_global = None
                    file1_sheet_name_global = None
                    file1_status.config(text="No sheets found", foreground="red")
                elif file_type == "Roaster":
                    file2_path_global = None
                    file2_sheet_name_global = None
                    file2_status.config(text="No sheets found", foreground="red")
                check_and_enable_merge_button()
                return

            if len(sheet_names) > 1:
                selected_sheet = ask_sheet_selection(sheet_names, file_path)
                if not selected_sheet:
                    # User cancelled sheet selection
                    if file_type == "Scheduler":
                        file1_status.config(text="Sheet selection cancelled", foreground="red")
                        file1_path_global = None
                        file1_sheet_name_global = None
                    elif file_type == "Roaster":
                        file2_status.config(text="Sheet selection cancelled", foreground="red")
                        file2_path_global = None
                        file2_sheet_name_global = None
                    check_and_enable_merge_button()
                    return
            else:
                selected_sheet = sheet_names[0]
            
            # Try to read the file with the selected sheet to catch immediate parsing errors
            # This is important to ensure the file is actually readable before setting global paths
            try:
                read_file_into_df(file_path, selected_sheet)
            except ValueError as ve:
                messagebox.showerror("File Read Error", f"Could not read selected Excel file or sheet: {ve}")
                if file_type == "Scheduler":
                    file1_path_global = None
                    file1_sheet_name_global = None
                    file1_status.config(text="Error reading sheet", foreground="red")
                elif file_type == "Roaster":
                    file2_path_global = None
                    file2_sheet_name_global = None
                    file2_status.config(text="Error reading sheet", foreground="red")
                check_and_enable_merge_button()
                return

            if file_type == "Scheduler":
                file1_path_global = file_path
                file1_sheet_name_global = selected_sheet
                file1_status.config(text=f"Selected: {os.path.basename(file1_path_global)} (Sheet: {file1_sheet_name_global})", foreground="green")
            elif file_type == "Roaster":
                file2_path_global = file_path
                file2_sheet_name_global = selected_sheet
                file2_status.config(text=f"Selected: {os.path.basename(file2_path_global)} (Sheet: {file2_sheet_name_global})", foreground="green")
        except ValueError as e:
            messagebox.showerror("File Error", str(e))
            if file_type == "Scheduler":
                file1_status.config(text="Error reading file", foreground="red")
                file1_path_global = None
                file1_sheet_name_global = None
            elif file_type == "Roaster":
                file2_path_global = None
                file2_sheet_name_global = None
    else: # CSV file - No sheet selection needed, but handle potential errors
        try:
            # Just try to read it to confirm it's valid
            read_file_into_df(file_path) 
            if file_type == "Scheduler":
                file1_path_global = file_path
                file1_sheet_name_global = None # No sheet for CSV
                file1_status.config(text=f"Selected: {os.path.basename(file1_path_global)}", foreground="green")
            elif file_type == "Roaster":
                file2_path_global = file_path
                file2_sheet_name_global = None # No sheet for CSV
                file2_status.config(text=f"Selected: {os.path.basename(file2_path_global)}", foreground="green")
        except Exception as e:
            messagebox.showerror("File Error", f"Error reading CSV file '{os.path.basename(file_path)}': {e}")
            if file_type == "Scheduler":
                file1_status.config(text="Error reading CSV", foreground="red")
                file1_path_global = None
                file1_sheet_name_global = None
            elif file_type == "Roaster":
                file2_path_global = None
                file2_sheet_name_global = None

    check_and_enable_merge_button()

def ask_sheet_selection(sheet_names, file_path):
    """Opens a new window for sheet selection."""
    dialog = tk.Toplevel(window)
    dialog.title("Select Sheet")
    dialog.geometry("300x150")
    dialog.transient(window)
    dialog.grab_set()

    tk.Label(dialog, text=f"Select a sheet from '{os.path.basename(file_path)}':", font=("Segoe UI", 10, "bold")).pack(pady=10)

    sheet_var = tk.StringVar(dialog)
    sheet_var.set(sheet_names[0]) # default value

    sheet_menu = ttk.Combobox(dialog, textvariable=sheet_var, values=sheet_names, state="readonly")
    sheet_menu.pack(pady=5)

    selected_sheet = None

    def on_ok():
        nonlocal selected_sheet
        selected_sheet = sheet_var.get()
        dialog.destroy()

    def on_cancel():
        nonlocal selected_sheet
        selected_sheet = None
        dialog.destroy()

    ok_button = ttk.Button(dialog, text="OK", command=on_ok)
    ok_button.pack(side="left", padx=(50, 10), pady=10)

    cancel_button = ttk.Button(dialog, text="Cancel", command=on_cancel)
    cancel_button.pack(side="right", padx=(10, 50), pady=10)

    window.wait_window(dialog) # Wait until the dialog is closed
    return selected_sheet

def check_and_enable_merge_button():
    if file1_path_global and file2_path_global:
        merge_button.config(state=tk.NORMAL)
        style.map('Merge.TButton',
                  background=[('disabled', 'lightgray'), ('!disabled', '#4CAF50')],
                  foreground=[('disabled', 'gray'), ('!disabled', 'white')])
        status_label.config(text="Ready to merge files", foreground="green")
    else:
        merge_button.config(state=tk.DISABLED)
        style.map('Merge.TButton',
                  background=[('disabled', 'lightgray'), ('!disabled', '#4CAF50')],
                  foreground=[('disabled', 'gray'), ('!disabled', 'white')])
        status_label.config(text="Please select both files to continue", foreground="orange")

def perform_merge_logic():
    global save_path_global, merged_df_global

    window.after(0, lambda: status_label.config(text="Processing merge...", foreground="blue"))
    window.after(0, lambda: merge_button.config(state=tk.DISABLED))
    window.after(0, lambda: select_file1_button.config(state=tk.DISABLED))
    window.after(0, lambda: select_file2_button.config(state=tk.DISABLED))
    window.after(0, lambda: export_button.config(state=tk.DISABLED)) # Disable export during merge
    window.after(0, lambda: exit_button.config(state=tk.DISABLED)) # Disable exit during merge

    try:
        # Read files using the enhanced read_file_into_df
        df1 = read_file_into_df(file1_path_global, file1_sheet_name_global)
        df2 = read_file_into_df(file2_path_global, file2_sheet_name_global)

        # Column stripping is now done inside read_file_into_df

        if 'NPI' not in df1.columns or 'VotedDate' not in df1.columns:
            raise ValueError("Scheduler file must contain 'NPI' and 'VotedDate' columns.")
        if 'Individual NPI' not in df2.columns or 'Provider Effective Date' not in df2.columns:
            raise ValueError("Roaster file must contain 'Individual NPI' and 'Provider Effective Date' columns.")

        # Date conversion is now done inside read_file_into_df, so no need here

        df2_original_dates = df2[['Individual NPI', 'Provider Effective Date']].copy()
        
        # Merge operation
        merged = pd.merge(df2, df1[['NPI', 'VotedDate']], how='left', left_on='Individual NPI', right_on='NPI')
        
        # Track if Provider Effective Date was originally NaT or None before update
        merged['Was_Originally_Empty'] = merged['Provider Effective Date'].isna()

        # Update 'Provider Effective Date' only if 'VotedDate' is not null
        merged['Provider Effective Date'] = merged['VotedDate'].fillna(merged['Provider Effective Date'])
        
        # Determine if 'Provider Effective Date' was updated
        merged['Was_Updated'] = False
        for idx, row in merged.iterrows():
            orig_date_series = df2_original_dates[df2_original_dates['Individual NPI'] == row['Individual NPI']]['Provider Effective Date']
            orig_date = orig_date_series.iloc[0] if not orig_date_series.empty else pd.NaT

            if pd.notna(row['VotedDate']): # If there is a VotedDate
                if pd.isna(orig_date): # If original Provider Effective Date was missing/NaT
                    merged.loc[idx, 'Was_Updated'] = True
                else:
                    # Compare only the date part, ignore time, for floating point precision issues
                    # Now that dates are already .dt.date, direct comparison works as intended
                    if row['VotedDate'] != orig_date: 
                        merged.loc[idx, 'Was_Updated'] = True

        merged_df_global = merged.copy() 

        # Ask for save path for the Excel output
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Merged_Output.xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not save_path:
            window.after(0, lambda: status_label.config(text="Merge cancelled", foreground="orange"))
            reset_gui()
            return

        save_path_global = save_path
        
        # Prepare DataFrame for direct Excel export (without Was_Updated, NPI, VotedDate columns)
        df_for_excel_output = merged_df_global.copy()
        if 'Was_Updated' in df_for_excel_output.columns:
            df_for_excel_output.drop(columns=['Was_Updated', 'NPI', 'VotedDate', 'Was_Originally_Empty'], inplace=True, errors='ignore')

        df_for_excel_output.to_excel(save_path, index=False)

        # Re-open the saved Excel file to apply conditional formatting
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active

        provider_idx = None
        for idx, cell in enumerate(ws[1], 1): # ws[1] is the first row (headers)
            if cell.value == "Provider Effective Date":
                provider_idx = idx
                break

        if provider_idx is None:
            raise ValueError("Could not find 'Provider Effective Date' column in the output for coloring.")

        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Apply coloring based on the original merged_df_global which still has 'Was_Updated'
        for df_idx, was_updated in enumerate(merged_df_global['Was_Updated']):
            if was_updated:
                excel_row = df_idx + 2 # +1 for 0-indexed to 1-indexed, +1 for header row
                ws.cell(row=excel_row, column=provider_idx).fill = fill
            
        wb.save(save_path)

        window.after(0, lambda: status_label.config(
            text=f"Merge completed successfully!\nSaved to: {os.path.basename(save_path_global)}", 
            foreground="green"))
        window.after(0, lambda: messagebox.showinfo(
            "Success", 
            f"File successfully saved at:\n{save_path_global}"))
        window.after(0, show_preview)
        window.after(0, lambda: export_button.config(state=tk.NORMAL)) # Enable export after successful merge

    except Exception as e:
        error_msg = str(e)
        window.after(0, lambda: status_label.config(text=f"Error: {error_msg}", foreground="red"))
        window.after(0, lambda: messagebox.showerror("Error", error_msg))
    finally:
        window.after(0, reset_gui)
        window.after(0, lambda: exit_button.config(state=tk.NORMAL)) # Enable exit button after merge operation

def start_merge_thread():
    threading.Thread(target=perform_merge_logic, daemon=True).start()

def reset_gui():
    merge_button.config(state=tk.DISABLED)
    select_file1_button.config(state=tk.NORMAL)
    select_file2_button.config(state=tk.NORMAL)
    # The export button should be enabled only if there is merged data available
    if merged_df_global is not None and not merged_df_global.empty:
        export_button.config(state=tk.NORMAL)
    else:
        export_button.config(state=tk.DISABLED)
    
    # Update button styles
    check_and_enable_merge_button() # Re-apply merge button style
    style.map('TButton',
              background=[('disabled', 'lightgray'), ('!disabled', '#E1E1E1')],
              foreground=[('disabled', 'gray'), ('!disabled', 'black')]) # Default button style

    if not (file1_path_global and file2_path_global):
        status_label.config(text="Please select both files to continue", foreground="orange")

def show_preview():
    global merged_df_global

    for w in preview_frame.winfo_children():
        w.destroy()

    if merged_df_global is None or merged_df_global.empty:
        return

    # Create a copy to display in preview, dropping 'Was_Updated', 'NPI', 'VotedDate', 'Was_Originally_Empty'
    df_for_preview = merged_df_global.copy()
    if 'Was_Updated' in df_for_preview.columns:
        df_for_preview.drop(columns=['Was_Updated', 'NPI', 'VotedDate', 'Was_Originally_Empty'], inplace=True, errors='ignore')

    cols = list(df_for_preview.columns)

    tree = ttk.Treeview(preview_frame, columns=cols, show="headings")
    vsb = ttk.Scrollbar(preview_frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(preview_frame, orient="horizontal", command=tree.xview)

    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    # Configure horizontal scroll increment for much faster movement
    hsb.config(command=lambda *args: tree.xview_scroll(int(args[1]) * 3, "units") if args[0] == "scroll" else tree.xview(*args)) # Tripled scroll speed
    tree.xscrollincrement = 90 # Adjusted value for desired speed (can be higher for more speed)

    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(expand=True, fill="both")

    tree.tag_configure("highlight", background="#FFF9C4") # Yellow highlight

    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=150, anchor="center")

    # Iterate through the original merged_df_global to get 'Was_Updated' status
    # and then use df_for_preview's values for display
    for i, row_orig in enumerate(merged_df_global.itertuples(index=False)):
        values_for_display = tuple(df_for_preview.iloc[i]) # Get values from the clean df for display
        
        was_updated = False
        if 'Was_Updated' in merged_df_global.columns:
            try:
                was_updated_col_index = merged_df_global.columns.get_loc('Was_Updated')
                was_updated = row_orig[was_updated_col_index]
            except IndexError:
                pass 

        tag = "highlight" if was_updated else ""
        tree.insert("", "end", values=values_for_display, tags=(tag,))

def export_data(file_format):
    global merged_df_global
    if merged_df_global is None or merged_df_global.empty:
        messagebox.showwarning("Export Warning", "No merged data available to export.")
        return

    # Prepare DataFrame for export: always remove internal 'NPI', 'VotedDate', 'Was_Originally_Empty' columns for final output
    df_to_export_clean = merged_df_global.copy()
    df_to_export_clean.drop(columns=['NPI', 'VotedDate', 'Was_Originally_Empty'], inplace=True, errors='ignore')

    if file_format == "Excel":
        # This "Export to Excel" here will just save a clean version without the Was_Updated column.
        if 'Was_Updated' in df_to_export_clean.columns:
            df_to_export_clean.drop(columns=['Was_Updated'], inplace=True, errors='ignore')

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Exported_Data.xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            try:
                df_to_export_clean.to_excel(file_path, index=False)
                messagebox.showinfo("Export Success", f"Data exported successfully to Excel:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

    elif file_format == "CSV":
        # For CSV, add a new column to indicate if 'Provider Effective Date' was updated
        df_for_csv_export = merged_df_global.copy()
        if 'Was_Updated' in df_for_csv_export.columns:
            # Map True/False to 'Yes'/'No' for better readability in CSV
            df_for_csv_export['Provider Effective Date Updated'] = df_for_csv_export['Was_Updated'].map({True: 'Yes', False: 'No'})
            # Drop the internal 'Was_Updated', 'NPI', 'VotedDate', 'Was_Originally_Empty' columns
            df_for_csv_export.drop(columns=['Was_Updated', 'NPI', 'VotedDate', 'Was_Originally_Empty'], inplace=True, errors='ignore')
        else:
            # If 'Was_Updated' column is somehow missing, just drop NPI and VotedDate
            df_for_csv_export.drop(columns=['NPI', 'VotedDate', 'Was_Originally_Empty'], inplace=True, errors='ignore')


        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile="Exported_Data.csv",
            filetypes=[("CSV Files", "*.csv")]
        )
        if file_path:
            try:
                df_for_csv_export.to_csv(file_path, index=False)
                messagebox.showinfo("Export Success", f"Data exported successfully to CSV:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export to CSV: {e}")

# ------------- Professional GUI Design -------------
window = tk.Tk()
window.title("Excel Data Merger - Schedule & Roaster")
window.state('zoomed') # Opens the window in maximized state by default
window.configure(bg="#f0f0f0")

# Load application icon (for taskbar/title bar)
script_dir = os.path.dirname(__file__)
app_icon_png_path = os.path.join(script_dir, 'Icons', 'app_icon.png') # Path to the PNG icon

if os.path.exists(app_icon_png_path):
    try:
        # Load the PNG image using PIL/Pillow
        icon_image_raw = Image.open(app_icon_png_path)
        icon_image = ImageTk.PhotoImage(icon_image_raw)
        window.iconphoto(True, icon_image) # Set the icon for the window
    except Exception as e:
        messagebox.showwarning("Icon Error", f"Could not set application icon from PNG: {e}. Ensure '{os.path.basename(app_icon_png_path)}' is a valid PNG file.")
else:
    messagebox.showwarning("Icon Warning", f"Application icon file '{os.path.basename(app_icon_png_path)}' not found. Taskbar icon may not appear.")

# Load internal button icons
try:
    schedule_icon_path = os.path.join(script_dir, 'Icons', 'scheduler_icon.png') 
    roaster_icon_path = os.path.join(script_dir, 'Icons', 'roaster_icon.png')
    
    schedule_icon_raw = Image.open(schedule_icon_path).resize((20, 20), Image.Resampling.LANCZOS)
    schedule_icon = ImageTk.PhotoImage(schedule_icon_raw)

    roaster_icon_raw = Image.open(roaster_icon_path).resize((20, 20), Image.Resampling.LANCZOS)
    roaster_icon = ImageTk.PhotoImage(roaster_icon_raw)

except FileNotFoundError:
    messagebox.showwarning("Icon Warning", "Could not load button icon files. Please ensure 'scheduler_icon.png' and 'roaster_icon.png' are in the 'Icons' subfolder next to the script.")
    schedule_icon = None
    roaster_icon = None
except Exception as e:
    messagebox.showwarning("Icon Error", f"Error loading button icons: {e}. Buttons will be text-only.")
    schedule_icon = None
    roaster_icon = None


# Custom style for buttons
style = ttk.Style()
style.theme_use('clam')

# Configure general ttk.Button style
style.configure('TButton', font=('Segoe UI', 10), padding=10, relief="groove")
style.map('TButton',
          background=[('disabled', 'lightgray'), ('!disabled', '#E1E1E1')],
          foreground=[('disabled', 'gray'), ('!disabled', 'black')])

# Configure specific style for the Merge button
style.configure('Merge.TButton', background="#4CAF50", foreground="white", font=('Segoe UI', 10, 'bold'), relief="groove")
style.map('Merge.TButton',
          background=[('disabled', 'lightgray'), ('!disabled', '#4CAF50')],
          foreground=[('disabled', 'gray'), ('!disabled', 'white')])


# Header
header = tk.Frame(window, bg="#0078D7", height=60)
header.pack(fill="x")

title = tk.Label(header, 
                 text="Excel Data Merger - Schedule & Roaster", 
                 font=("Segoe UI", 16, "bold"), 
                 bg="#0078D7", 
                 fg="white")
title.pack(pady=15)

# Main content
content = tk.Frame(window, bg="#f0f0f0")
content.pack(expand=True, fill="both", padx=20, pady=10)

# File selection panel
file_panel = tk.LabelFrame(content, 
                            text=" File Selection ",
                            font=("Segoe UI", 11, "bold"),
                            bg="#f0f0f0",
                            padx=10,
                            pady=10)
file_panel.pack(fill="x", pady=(0, 15))

# Frame for file selection buttons (left side of file_panel)
file_selection_buttons_frame = tk.Frame(file_panel, bg="#f0f0f0")
file_selection_buttons_frame.pack(side="left", fill="y", padx=(0, 20)) # Pack to the left

# File 1 selection
file1_frame = tk.Frame(file_selection_buttons_frame, bg="#f0f0f0")
file1_frame.pack(fill="x", pady=5)

select_file1_button = ttk.Button(file1_frame, 
                                 text="Import Scheduler File",
                                 command=lambda: select_file_and_sheet("Scheduler"),
                                 compound="left", # Place icon to the left of text
                                 image=schedule_icon,
                                 width=20) # Increased width
select_file1_button.pack(side="left", padx=(0, 10))

file1_status = tk.Label(file1_frame, 
                         text="No file selected", 
                         bg="#f0f0f0", 
                         font=("Segoe UI", 10))
file1_status.pack(side="left")

# File 2 selection
file2_frame = tk.Frame(file_selection_buttons_frame, bg="#f0f0f0")
file2_frame.pack(fill="x", pady=5)

select_file2_button = ttk.Button(file2_frame, 
                                 text="Import Roaster File",
                                 command=lambda: select_file_and_sheet("Roaster"),
                                 compound="left", # Place icon to the left of text
                                 image=roaster_icon,
                                 width=20) # Increased width
select_file2_button.pack(side="left", padx=(0, 10))

file2_status = tk.Label(file2_frame, 
                         text="No file selected", 
                         bg="#f0f0f0", 
                         font=("Segoe UI", 10))
file2_status.pack(side="left")

# Action buttons - Moved to the right top corner of the file_panel
# Create a frame to hold these buttons and pack it to the right
action_buttons_frame = tk.Frame(file_panel, bg="#f0f0f0")
action_buttons_frame.pack(side="right", fill="y", anchor="ne", padx=(0,10), pady=(0,10)) # Anchor top-right

# Pack buttons from right to left within action_buttons_frame to achieve "Merge, Export, Exit" sequence from left to right on the GUI
exit_button = ttk.Button(action_buttons_frame, 
                         text="Exit", 
                         command=window.destroy,
                         width=10) # Adjusted width
exit_button.pack(side="right", padx=(10, 0), pady=5) # Pack to the right

export_button = ttk.Menubutton(action_buttons_frame, text="Export", state=tk.DISABLED, direction="below", width=15)
export_button.pack(side="right", padx=(10, 0), pady=5) # Pack to the right

export_menu = tk.Menu(export_button, tearoff=0)
export_menu.add_command(label="Export to Excel", command=lambda: export_data("Excel"))
export_menu.add_command(label="Export to CSV", command=lambda: export_data("CSV"))
export_button["menu"] = export_menu

merge_button = ttk.Button(action_buttons_frame, 
                          text="Merge Files", 
                          state=tk.DISABLED,
                          command=start_merge_thread,
                          style='Merge.TButton',
                          width=15) # Adjusted width
merge_button.pack(side="right", padx=(10, 0), pady=5) # Pack to the right

# Status bar (remains below the file_panel)
status_frame = tk.Frame(content, bg="#f0f0f0")
status_frame.pack(fill="x", pady=(0, 15))

status_label = tk.Label(status_frame, 
                         text="Please select both files to continue", 
                         bg="#f0f0f0", 
                         font=("Segoe UI", 10),
                         fg="orange")
status_label.pack()

# Preview panel
preview_panel = tk.LabelFrame(content, 
                               text=" Merged Data Preview ",
                               font=("Segoe UI", 11, "bold"),
                               bg="#f0f0f0",
                               padx=10,
                               pady=10)
preview_panel.pack(expand=True, fill="both")

preview_frame = tk.Frame(preview_panel, bg="#f0f0f0")
preview_frame.pack(expand=True, fill="both", padx=5, pady=5)

# Initial check to set button states
check_and_enable_merge_button()

window.mainloop()
